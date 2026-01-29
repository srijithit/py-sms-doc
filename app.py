from flask import Flask, request, redirect, render_template_string, session, flash
from datetime import datetime
import random, os, time
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "secret123"

# ================== LANGUAGE LABELS ==================
LANG = {
    "en": {
        "title": "Telemedicine Portal",
        "patient_login": "Patient Login",
        "patient_register": "Patient Register",
        "doctor_login": "Doctor Login",
        "username": "Username",
        "password": "Password",
        "login": "Login",
        "register": "Register",
        "symptoms": "Describe symptoms",
        "send": "Send",
        "logout": "Logout",
        "welcome": "Welcome",
        "doctor_dashboard": "Doctor Dashboard",
        "submit": "Submit",
        "submitted": "Submitted. AI advice ready.",
        "ai_sms": "🤖 AI SMS",
        "sms_doctor": "📩 SMS Doctor",
        "login_title": "Patient Login",
"register_title": "Patient Registration",
"create_username": "Create Username",
"create_password": "Create Password",
"full_name": "Full Name",
"age": "Age",
"gender": "Gender",
"condition": "Condition",
"area": "Area / Village",
"phone": "Phone",
"back_home": "⬅ Home",
"doctor_pin": "Doctor PIN"


    },
    "ta": {
        "title": "தொலை மருத்துவ சேவை",
        "patient_login": "நோயாளர் உள்நுழைவு",
        "patient_register": "நோயாளர் பதிவு",
        "doctor_login": "மருத்துவர் உள்நுழைவு",
        "username": "பயனர் பெயர்",
        "password": "கடவுச்சொல்",
        "submitted": "சமர்ப்பிக்கப்பட்டது. AI ஆலோசனை தயாராக உள்ளது.",    
        "login": "உள்நுழை",
        "register": "பதிவு",
        "symptoms": "அறிகுறிகளை எழுதவும்",
        "send": "அனுப்பு",
        "logout": "வெளியேறு",
        "welcome": "வரவேற்பு",
        "doctor_dashboard": "மருத்துவர் கட்டுப்பாடு",
        "submit": "சமர்ப்பிக்கவும்",
        "ai_sms": "🤖 AI எஸ்எம்எஸ்",
"sms_doctor": "📩 மருத்துவருக்கு எஸ்எம்எஸ்",
"login_title": "நோயாளர் உள்நுழைவு",
"register_title": "நோயாளர் பதிவு",
"create_username": "பயனர் பெயரை உருவாக்கவும்",
"create_password": "கடவுச்சொல் உருவாக்கவும்",
"full_name": "முழு பெயர்",
"age": "வயது",
"gender": "பாலினம்",
"condition": "நோய் நிலை",
"area": "பகுதி / கிராமம்",
"phone": "தொலைபேசி",
"back_home": "⬅ முகப்பு",
"doctor_pin": "மருத்துவர் பின்"



    },
    "hi": {
        "title": "टेलीमेडिसिन पोर्टल",
        "patient_login": "मरीज़ लॉगिन",
        "patient_register": "मरीज़ पंजीकरण",
        "doctor_login": "डॉक्टर लॉगिन",
        "username": "उपयोगकर्ता नाम",
        "password": "पासवर्ड",
        "login": "लॉगिन",
        "submitted": "सबमिट किया गया। एआई सलाह तैयार है।",
        "register": "पंजीकरण",
        "symptoms": "लक्षण लिखें",
        "send": "भेजें",
        "logout": "लॉगआउट",
        "welcome": "स्वागत है",
        "doctor_dashboard": "डॉक्टर डैशबोर्ड",
        "submit": "जमा करें",
        "ai_sms": "🤖 एआई एसएमएस",
        "sms_doctor": "📩 डॉक्टर को एसएमएस",
        "login_title": "मरीज़ लॉगिन",
"register_title": "मरीज़ पंजीकरण",
"create_username": "यूज़रनेम बनाएं",
"create_password": "पासवर्ड बनाएं",
"full_name": "पूरा नाम",
"age": "उम्र",
"gender": "लिंग",
"condition": "बीमारी",
"area": "क्षेत्र / गाँव",
"phone": "फ़ोन",
"back_home": "⬅ होम",
"doctor_pin": "डॉक्टर पिन"



    }
}

@app.before_request
def set_lang():
    if "lang" not in session:
        session["lang"] = "en"
    if request.args.get("lang") in ["en","ta","hi"]:
        session["lang"] = request.args.get("lang")

def t(key):
    return LANG.get(session.get("lang","en"), LANG["en"]).get(key, key)

# ✅ THIS LINE IS REQUIRED
app.jinja_env.globals.update(t=t)


# ================== DATA ==================
os.makedirs("data", exist_ok=True)
EXCEL = "data/hospital_data.xlsx"
DOCTOR_PIN = "7788"
DOCTOR_PHONE = "9791471277"   # change if needed

# ================== EXCEL INIT ==================
def init_excel():
    if not os.path.exists(EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"
        ws.append(["Username","Name","Age","Gender","Condition","Area","Phone"])
        wb.create_sheet("Appointments").append(["Username","Symptoms","Time"])
        wb.create_sheet("Replies").append(
            ["Username","Reply","Time","Medicine","Dose","Days"]
        )
        wb.save(EXCEL)
        wb.close()

init_excel()

# ================== MEMORY ==================
users = {"patient":{}}
appointments = []
prescriptions = []

# ================== HELPERS ==================
def sms_link(phone, msg):
    msg = msg.replace("\n","%0A").replace(" ","%20")
    return f"sms:{phone}?body={msg}"

def safe_save(func,*a):
    for _ in range(3):
        try:
            func(*a)
            return
        except PermissionError:
            time.sleep(0.5)

def save_patient(u,f):
    wb = load_workbook(EXCEL)
    wb["Patients"].append([
        u,f["name"],f["age"],f["gender"],
        f["condition"],f["area"],f["phone"]
    ])
    wb.save(EXCEL); wb.close()

def save_appointment(u,s):
    wb = load_workbook(EXCEL)
    wb["Appointments"].append([u,s,datetime.now()])
    wb.save(EXCEL); wb.close()

def save_reply(u,r):
    wb = load_workbook(EXCEL)
    wb["Replies"].append([u,r,datetime.now()])
    wb.save(EXCEL); wb.close()

# ================== AI AUTO REPLY ==================
def ai_auto_reply(symptoms):
    s = symptoms.lower()
    if any(x in s for x in ["fever","temperature","hot"]):
        return "AI Advice:\nFever suspected.\nDrink fluids, paracetamol.\nSee doctor if >2 days."
    if any(x in s for x in ["cough","cold","throat"]):
        return "AI Advice:\nCold/Cough.\nSteam, warm water.\nAvoid cold food."
    if any(x in s for x in ["vomit","diarrhea","loose"]):
        return "AI Advice:\nStomach issue.\nORS + light food."
    if any(x in s for x in ["chest","breath","heart"]):
        return "🚨 EMERGENCY 🚨\nBreathing/Chest pain.\nCALL 108 NOW."
    return "AI Advice:\nRest & monitor.\nDoctor will reply soon."

# ================== UI ==================
STYLE = """
<style>
body{font-family:Arial;min-height:100vh;margin:0;
display:flex;justify-content:center;align-items:center;
background:linear-gradient(rgba(0,0,0,.4),rgba(0,0,0,.4)),
url('https://images.unsplash.com/photo-1586773860418-d37222d8fce3');
background-size:cover}
.box{background:white;padding:20px;border-radius:14px;
width:95%;max-width:420px;text-align:center;
box-shadow:0 10px 30px rgba(0,0,0,.3)}
input,textarea,button{width:100%;padding:12px;margin:8px 0}
button{background:#007bff;color:white;border:none;border-radius:6px}
.card{border:1px solid #ddd;border-radius:8px;padding:10px;margin-top:10px}
table{width:100%;border-collapse:collapse}
th,td{border:1px solid #ccc;padding:6px}
</style>
"""

HOME = STYLE + """

<div class=box>
<select onchange="location='?lang='+this.value">
  <option value="en">English</option>
  <option value="ta">தமிழ்</option>
  <option value="hi">हिंदी</option>
</select>
<h2>{{ t('title') }}</h2>
<a href="/login"><button style="background:#00a1ff">{{ t('patient_login') }}</button></a>
<a href="/register"><button style="background:#28a745">{{ t('patient_register') }}</button></a>
<form method=post action="/doctor-pin">
<input type="password" name="pin" placeholder="{{ t('doctor_pin') }}">
<button style="background:#6c757d">{{ t('doctor_login') }}</button>
<button type="button" onclick="sendSOS()" style="background:#dc3545">🚑 SOS 108</button>
</form>
<script>
function sendSOS(){
 window.location.href="tel:108";
}
</script>
</div>
"""

LOGIN = STYLE + """

<div class=box>
<select onchange="location='?lang='+this.value">
  <option value="en">English</option>
  <option value="ta">தமிழ்</option>
  <option value="hi">हिंदी</option>
</select>

<h3>{{ t('login_title') }}</h3>

<input name=username placeholder="{{ t('username') }}" required>
<input type=password name=password placeholder="{{ t('password') }}" required>

<button>{{ t('login') }}</button>

<a href="/">{{ t('back_home') }}</a>



</div>
"""

REGISTER = STYLE + """


<div class=box>
<select onchange="location='?lang='+this.value">
  <option value="en">English</option>
  <option value="ta">தமிழ்</option>
  <option value="hi">हिंदी</option>
</select>
<h3>{{ t('register_title') }}</h3>

<input name=username placeholder="{{ t('create_username') }}" required>
<input type=password name=password placeholder="{{ t('create_password') }}" required>
<input name=name placeholder="{{ t('full_name') }}" required>
<input name=age placeholder="{{ t('age') }}" required>
<input name=gender placeholder="{{ t('gender') }}" required>
<input name=condition placeholder="{{ t('condition') }}" required>
<input name=area placeholder="{{ t('area') }}" required>
<input name=phone placeholder="{{ t('phone') }}" required>

<button>{{ t('register') }}</button>

<a href="/">{{ t('back_home') }}</a>

</div>
"""

PATIENT = STYLE + """


<div class=box>
<select onchange="location='?lang='+this.value">
  <option value="en">English</option>
  <option value="ta">தமிழ்</option>
  <option value="hi">हिंदी</option>
</select>
{% with m=get_flashed_messages() %}
{% if m %}<script>alert("{{m[0]}}");</script>{% endif %}
{% endwith %}
<h3>{{ t('welcome') }} {{ user }}</h3>
<textarea name=symptoms placeholder="{{ t('symptoms') }}"></textarea>
<button>{{ t('send') }}</button>


{% if session.get('sms_patient') %}
<a href="{{session.pop('sms_patient')}}">
<button>{{ t('ai_sms') }}</button>
{% endif %}

{% if session.get('sms_doctor') %}
<a href="{{session.pop('sms_doctor')}}">
<button>{{ t('sms_doctor') }}</button>
{% endif %}

{% for p in data %}
<div class=card>
<b>Date:</b>{{p.date}}<br>
<b>Symptoms:</b>{{p.symptoms}}<br>
<b>Advice:</b>{{p.reply}}
</div>
{% endfor %}
<a href=/logout>Logout</a>
</div>
"""

DOCTOR = STYLE + """


<div class=box>
<select onchange="location='?lang='+this.value">
  <option value="en">English</option>
  <option value="ta">தமிழ்</option>
  <option value="hi">हिंदी</option>
</select>
<h3>{{ t('doctor_dashboard') }}</h3>
{% for a in apps %}
<form method=post action=/reply>
<b>{{a.patient}}</b><br>{{a.symptoms}}
<textarea name=reply required></textarea>
<input name=m1 placeholder="Medicine">
<input name=d1 placeholder="Dose">
<input name=days1 placeholder="Days">
<input type=hidden name=patient value="{{a.patient}}">
<button>Submit</button>
</form><hr>
{% endfor %}
{% if sms %}
<a href="{{sms}}"><button style="background:#28a745">📩 SMS Patient</button></a>
{% endif %}
<a href=/logout>Logout</a>
</div>
"""

# ================== ROUTES ==================
@app.route("/")
def home(): return render_template_string(HOME)

@app.route("/login",methods=["GET","POST"])
def login():
    if request.method=="POST":
        u=request.form["username"]
        if users["patient"].get(u)==request.form["password"]:
            session["user"]=u; session["role"]="patient"
            return redirect("/dashboard")
    return render_template_string(LOGIN)

@app.route("/doctor-pin",methods=["POST"])
def doctor_pin():
    if request.form["pin"]==DOCTOR_PIN:
        session["role"]="doctor"
        return redirect("/dashboard")
    return redirect("/")

@app.route("/register",methods=["GET","POST"])
def register():
    msg=""
    if request.method=="POST":
        u=request.form["username"]
        if u in users["patient"]:
            msg="❌ Username already exists"
        else:
            users["patient"][u]=request.form["password"]
            users[u+"_phone"]=request.form["phone"]
            users[u+"_area"]=request.form["area"]
            safe_save(save_patient,u,request.form)
            msg="✅ Registration successful. Login now."
    return render_template_string(REGISTER,msg=msg)

@app.route("/dashboard")
def dashboard():
    if session["role"]=="patient":
        return render_template_string(
            PATIENT,user=session["user"],
            data=[p for p in prescriptions if p["patient"]==session["user"]]
        )
    return render_template_string(DOCTOR,apps=appointments,sms=session.pop("sms",None))

@app.route("/submit",methods=["POST"])
def submit():
    u=session["user"]; s=request.form["symptoms"]
    appointments.append({"patient":u,"symptoms":s})
    safe_save(save_appointment,u,s)

    session["sms_patient"]=sms_link(users[u+"_phone"],ai_auto_reply(s))
    session["sms_doctor"]=sms_link(DOCTOR_PHONE,f"Patient:{u}\nSymptoms:{s}")

    flash(t("submitted"))
    return redirect("/dashboard")

@app.route("/reply",methods=["POST"])
def reply():
    u = request.form["patient"]

    sym = ""
    for i,a in enumerate(appointments):
        if a["patient"] == u:
            sym = a["symptoms"]
            appointments.pop(i)
            break

    prescriptions.append({
        "patient": u,
        "date": datetime.now().strftime("%d-%m-%Y"),
        "symptoms": sym,
        "reply": request.form["reply"],
        "medicine": request.form["m1"],
        "dose": request.form["d1"],
        "days": request.form["days1"]
    })

    sms_msg = (
        "DOCTOR PRESCRIPTION\n"
        f"Advice: {request.form['reply']}\n"
        f"Medicine: {request.form['m1']}\n"
        f"Dose: {request.form['d1']}\n"
        f"Days: {request.form['days1']}"
    )

    session["sms"] = sms_link(users[u+"_phone"], sms_msg)
    return redirect("/dashboard")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ================== RUN ==================
if __name__=="__main__":
    app.run(host="0.0.0.0",port=8000,debug=True)
